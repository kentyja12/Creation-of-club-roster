import pandas as pd
import os
import datetime
from openpyxl import load_workbook
import tkinter

class default:
    #名簿を書き込む原本の読み込み
    load_meibo_path = os.path.join(os.path.dirname(__file__),"resources\genpon.xlsx") #原本のパス取得
    wb = load_workbook(load_meibo_path) #原本の読み込み
    ws = wb.active 

    Columns=["学籍番号","氏名","役職","電話番号"] 
    now_date = datetime.datetime.now() 
    yyyymmdd = "{:04}{:02}{:02}".format(now_date.year,now_date.month,now_date.day)
    now_date = datetime.datetime.now()
    limit_count = 30 #登録人数の上限

class tmp:
    EditBox_tmp = '' #学籍番号
    EditBox_instant1_tmp = '' #インスタント学籍番号
    EditBox_instant2_tmp = '' #インスタント名前
    EditBox_instant3_tmp = '' #インスタント電話番号

class student_data:
    club_name = 'club_name' #クラブ名
    df_meibo = '' #読み込んだ名簿
    gakuban_tmp = '' #名簿入力欄に入力した学籍番号
    gakuban_list = [] #読み込んだ名簿から登録された学籍番号
    count = 0 #名簿に登録できる人数をカウント
    df = pd.DataFrame(columns = default.Columns) #インスタント名簿から登録された学生情報

    default.ws['A4'] = '●活動日：　{} 月　{} 日'.format(default.now_date.month,default.now_date.day) #日付入力
